"""One-off cleaner for experiment_data.xlsx (phone/group types, dup group 023, slot_map, meta)."""
from __future__ import annotations

import json
import os
import sys

from openpyxl import load_workbook

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_FILE = os.path.join(BASE, "data", "experiment_data.xlsx")

PHONE_COLS = {"phone", "suspect_phone", "interviewer_phone"}
GROUP_COLS = {"group_name", "name"}


def _parse_group_number(group_name):
    s = str(group_name or "").strip()
    if not s.isdigit():
        return 0
    try:
        n = int(s)
    except ValueError:
        return 0
    return n if 1 <= n <= 999 else 0


def _rows(wb, name):
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            out.append(dict(zip(headers, row)))
    return out, headers


def _max_id(rows):
    vals = []
    for r in rows:
        try:
            vals.append(int(r.get("id")))
        except (TypeError, ValueError):
            pass
    return max(vals) if vals else 0


def _normalize_cell_types(wb):
    counts = {"phone": 0, "group": 0}
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2):
            for i, col in enumerate(headers):
                if col is None:
                    continue
                c = row[i]
                v = c.value
                if col in PHONE_COLS:
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        c.value = str(int(v)) if float(v).is_integer() else str(v)
                        c.number_format = "@"
                        counts["phone"] += 1
                    elif isinstance(v, str):
                        c.number_format = "@"
                elif col in GROUP_COLS:
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        c.value = f"{int(v):03d}"
                        c.number_format = "@"
                        counts["group"] += 1
                    elif isinstance(v, str) and v.strip():
                        c.number_format = "@"
    return counts


def _fix_duplicate_group_023(wb):
    """Restore original 023 pair; move wrongly assigned suspect to next free group."""
    participants, _ = _rows(wb, "participants")
    groups, _ = _rows(wb, "groups")

    original_suspect = "18940216330"
    original_interviewer = "17502332345"
    wrong_suspect = "13002268524"

    used = set()
    for p in participants:
        n = _parse_group_number(p.get("group_name"))
        if n:
            used.add(n)

    new_group = None
    for n in range(1, 1000):
        if n not in used:
            new_group = f"{n:03d}"
            break
    if not new_group:
        raise RuntimeError("No free group number available")

    ws_p = wb["participants"]
    headers_p = [c.value for c in ws_p[1]]
    idx = {h: i for i, h in enumerate(headers_p)}

    for row in ws_p.iter_rows(min_row=2):
        phone = str(row[idx["phone"]].value or "").strip()
        if phone == wrong_suspect:
            row[idx["group_name"]].value = new_group
            fid = str(row[idx["full_id"]].value or "")
            if fid.startswith("023-"):
                row[idx["full_id"]].value = fid.replace("023-", f"{new_group}-", 1)
            row[idx["group_name"]].number_format = "@"
            if row[idx["full_id"]].value:
                row[idx["full_id"]].number_format = "@"

    ws_g = wb["groups"]
    headers_g = [c.value for c in ws_g[1]]
    gidx = {h: i for i, h in enumerate(headers_g)}
    found_023 = False
    found_new = False
    for row in ws_g.iter_rows(min_row=2):
        name = str(row[gidx["name"]].value or "").strip()
        if name == "023":
            row[gidx["suspect_phone"]].value = original_suspect
            row[gidx["interviewer_phone"]].value = original_interviewer
            found_023 = True
        if name == new_group:
            row[gidx["suspect_phone"]].value = wrong_suspect
            row[gidx["interviewer_phone"]].value = ""
            found_new = True
    if not found_023:
        ws_g.append(["023", original_suspect, original_interviewer, ""])
    if not found_new:
        ws_g.append([new_group, wrong_suspect, "", ""])

    return new_group


def _rebuild_slot_map(wb):
    participants, _ = _rows(wb, "participants")
    appointments, _ = _rows(wb, "appointments")
    pby = {str(p.get("phone") or "").strip(): p for p in participants}
    slot_map = {}
    for appt in appointments:
        if appt.get("status") != "confirmed":
            continue
        slot = str(appt.get("time_slot") or "").strip()
        phone = str(appt.get("phone") or "").strip()
        p = pby.get(phone)
        gn = str(p.get("group_name") or "").strip() if p else ""
        if not slot or not gn:
            continue
        if slot in slot_map and slot_map[slot] != gn:
            continue
        slot_map[slot] = gn
    return slot_map


def _set_meta(wb, key, value):
    ws = wb["meta"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == key:
            row[1].value = value
            return
    ws.append([key, value])


def clean(path=DEFAULT_FILE):
    wb = load_workbook(path)
    type_counts = _normalize_cell_types(wb)
    new_group = _fix_duplicate_group_023(wb)
    slot_map = _rebuild_slot_map(wb)
    _set_meta(wb, "appointment_slot_groups", json.dumps(slot_map, ensure_ascii=False))

    appts, _ = _rows(wb, "appointments")
    training, _ = _rows(wb, "training_sessions")
    qs, _ = _rows(wb, "interview_questionnaires")
    qo, _ = _rows(wb, "questionnaire_overrides")
    av, _ = _rows(wb, "availabilities")

    _set_meta(wb, "next_avail_id", _max_id(av) + 1 or 1)
    _set_meta(wb, "next_appt_id", _max_id(appts) + 1 or 1)
    _set_meta(wb, "next_training_session_id", _max_id(training) + 1 or 1)
    _set_meta(wb, "next_questionnaire_id", _max_id(qs) + 1 or 1)
    _set_meta(wb, "next_qoverride_id", _max_id(qo) + 1 or 1)

    wb.save(path)
    wb.close()
    return {
        "phone_cells_fixed": type_counts["phone"],
        "group_cells_fixed": type_counts["group"],
        "moved_wrong_023_to": new_group,
        "slot_map_entries": len(slot_map),
    }


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    result = clean(target)
    print("Clean complete:", result)
