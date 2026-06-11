import sys
from openpyxl import load_workbook

def main():
    out_path = sys.argv[1]
    wb = load_workbook(out_path)
    
    # Rename sheets to correct names
    sheet_mapping = {
        "Participants": "participants",
        "Appointments": "groups",
        "AppointmentSlots": "profiles",
        "Questionnaires": "availabilities",
        "QuestionnaireAnswers": "appointments",
        "TrainingSessions": "training_sessions"
    }
    
    for old_name, new_name in sheet_mapping.items():
        if old_name in wb.sheetnames:
            wb[old_name].title = new_name
            
    # Add missing sheets
    missing_sheets = {
        "interview_questionnaires": ["id", "phone", "role", "phase", "appointment_slot", "answers_json", "submitted_at"],
        "questionnaire_overrides": ["id", "phone", "phase", "is_open", "updated_at"],
        "serious_game_choices": ["timestamp", "phone", "participant_id", "case", "condition", "step_index", "video", "choice"],
        "meta": ["key", "value"]
    }
    
    for name, headers in missing_sheets.items():
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(headers)
            print(f"Added missing sheet: {name}")
            
    # Ensure correct columns for existing sheets
    SHEET_COLUMNS = {
        "participants": [
            "id", "phone", "role", "group_name", "full_id", "guilt", "case_type", "training_type",
            "consent_attention_passed", "attention_passed", "attention_failed",
            "sue_attention_passed", "sue_attention_attempts", "control_attention_passed",
            "game_completed", "profile_completed", "completed", "flow_step",
            "case_evidence_recap_passed",
            "created_at", "avatar_practice_transcript",
            "training_avatar_order", "training_ui_order",
        ],
        "groups": ["name", "suspect_id", "interviewer_id", "created_at"],
        "profiles": ["participant_id", "data", "submitted_at"],
        "availabilities": ["id", "phone", "group_name", "role", "slots", "updated_at"],
        "appointments": ["id", "phone", "time_slot", "role", "status", "booked_at"],
        "training_sessions": ["id", "interviewer_id", "phone", "session_num", "avatar_setting", "avatar_guilt", "judgment", "transcript", "feedback", "started_at", "completed_at"],
    }
    
    for name, headers in SHEET_COLUMNS.items():
        if name in wb.sheetnames:
            ws = wb[name]
            # Replace the first row with the correct headers, just in case
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)
                
    wb.save(out_path)
    print(f"Finalized recovered workbook: {out_path}")

if __name__ == '__main__':
    main()
