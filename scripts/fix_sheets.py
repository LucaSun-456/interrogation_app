from openpyxl import load_workbook
import sys

wb = load_workbook(r'e:\Working Documents\【Important】My Projects\24 Specific avatar\interrogation_app\data\experiment_data.xlsx')
print("Before:", wb.sheetnames)

if 'participants' in wb.sheetnames and 'participants1' in wb.sheetnames:
    del wb['participants']
    wb['participants1'].title = 'participants'

print("After:", wb.sheetnames)
wb.save(r'e:\Working Documents\【Important】My Projects\24 Specific avatar\interrogation_app\data\experiment_data.xlsx')
