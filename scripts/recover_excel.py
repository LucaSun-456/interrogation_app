#!/usr/bin/env python3
"""
Recover data from a damaged experiment_data.xlsx (or .corrupt.* backup).

Usage:
  python scripts/recover_excel.py path/to/experiment_data.xlsx.corrupt.20260611_174451
  python scripts/recover_excel.py corrupt.xlsx -o data/experiment_data_recovered.xlsx

The script tries, in order:
  1. openpyxl normal load
  2. openpyxl read_only + keep_links=False
  3. Zip member extract + XML sheet parse (works when zip is partially readable)
  4. Repack surviving zip members into a new .xlsx shell
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

# Run from project root
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import Workbook, load_workbook

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def zip_health(path: str) -> dict:
    info = {"size": os.path.getsize(path), "valid_zip": False, "members": [], "error": None}
    try:
        with zipfile.ZipFile(path, "r") as zf:
            info["valid_zip"] = True
            info["members"] = zf.namelist()
    except Exception as e:
        info["error"] = str(e)
    return info


def try_openpyxl(path: str, read_only: bool = False):
    kwargs = {"read_only": read_only, "data_only": True}
    if read_only:
        kwargs["keep_links"] = False
    wb = load_workbook(path, **kwargs)
    return wb


def _col_letters_to_index(col: str) -> int:
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _parse_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    path = "xl/sharedStrings.xml"
    if path not in zf.namelist():
        return []
    root = ET.fromstring(zf.read(path))
    out = []
    for si in root.findall("m:si", NS):
        parts = []
        for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"):
            if t.text:
                parts.append(t.text)
        out.append("".join(parts))
    return out


def _sheet_xml_to_rows(zf: zipfile.ZipFile, sheet_path: str, shared: list[str]) -> list[list]:
    root = ET.fromstring(zf.read(sheet_path))
    rows_map: dict[int, dict[int, str]] = {}
    for row in root.findall(".//m:sheetData/m:row", NS):
        r_idx = int(row.attrib.get("r", "0"))
        rows_map.setdefault(r_idx, {})
        for c in row.findall("m:c", NS):
            ref = c.attrib.get("r", "")
            m = re.match(r"^([A-Z]+)(\d+)$", ref)
            if not m:
                continue
            col_idx = _col_letters_to_index(m.group(1))
            cell_type = c.attrib.get("t")
            v = c.find("m:v", NS)
            is_elem = c.find("m:is", NS)
            val = ""
            if cell_type == "s" and v is not None and v.text is not None:
                idx = int(v.text)
                val = shared[idx] if 0 <= idx < len(shared) else v.text
            elif cell_type == "inlineStr" and is_elem is not None:
                t = is_elem.find(".//m:t", NS)
                val = t.text if t is not None and t.text else ""
            elif v is not None and v.text is not None:
                val = v.text
            rows_map[r_idx][col_idx] = val
    if not rows_map:
        return []
    max_row = max(rows_map)
    max_col = max(max(cols) for cols in rows_map.values())
    table = []
    for r in range(1, max_row + 1):
        line = []
        cols = rows_map.get(r, {})
        for c in range(1, max_col + 1):
            line.append(cols.get(c, ""))
        table.append(line)
    return table


def recover_via_xml(path: str) -> Workbook | None:
    try:
        zf = zipfile.ZipFile(path, "r")
    except zipfile.BadZipFile:
        return None
    shared = _parse_shared_strings(zf)
    sheet_files = sorted(
        n for n in zf.namelist()
        if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")
    )
    if not sheet_files:
        zf.close()
        return None
    # Sheet names from workbook.xml if possible
    names: list[str] = []
    if "xl/workbook.xml" in zf.namelist():
        try:
            wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
            for sh in wb_root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
                names.append(sh.attrib.get("name", f"sheet{len(names)+1}"))
        except ET.ParseError:
            pass
    wb = Workbook()
    wb.remove(wb.active)
    for i, sf in enumerate(sheet_files):
        try:
            rows = _sheet_xml_to_rows(zf, sf, shared)
        except ET.ParseError:
            continue
        title = names[i] if i < len(names) else f"recovered_{i+1}"
        ws = wb.create_sheet(title[:31])
        for row in rows:
            ws.append(row)
    zf.close()
    if len(wb.sheetnames) == 0:
        return None
    return wb


def repack_zip(path: str, out_path: str) -> bool:
    """Copy all readable zip members into a fresh xlsx container."""
    try:
        zf = zipfile.ZipFile(path, "r")
    except zipfile.BadZipFile:
        return False
    ok = 0
    with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as out:
        for name in zf.namelist():
            try:
                data = zf.read(name)
            except Exception:
                continue
            out.writestr(name, data)
            ok += 1
    zf.close()
    return ok > 3


def copy_workbook(wb, out_path: str) -> None:
    if hasattr(wb, "save"):
        wb.save(out_path)
        if hasattr(wb, "close"):
            wb.close()
        return
    # read_only workbook
    new_wb = Workbook()
    new_wb.remove(new_wb.active)
    for name in wb.sheetnames:
        src = wb[name]
        dst = new_wb.create_sheet(name[:31])
        for row in src.iter_rows(values_only=True):
            dst.append(list(row))
    wb.close()
    new_wb.save(out_path)
    new_wb.close()


def main():
    parser = argparse.ArgumentParser(description="Recover damaged experiment_data.xlsx")
    parser.add_argument("input", help="Path to corrupt .xlsx file")
    parser.add_argument(
        "-o", "--output",
        default="",
        help="Output path (default: experiment_data_recovered_<timestamp>.xlsx)",
    )
    args = parser.parse_args()
    src = os.path.abspath(args.input)
    if not os.path.isfile(src):
        print(f"ERROR: file not found: {src}", file=sys.stderr)
        sys.exit(1)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = args.output or os.path.join(
        os.path.dirname(src),
        f"experiment_data_recovered_{stamp}.xlsx",
    )

    print("=== Diagnosis ===")
    health = zip_health(src)
    print(f"  Size: {health['size']} bytes")
    print(f"  Valid ZIP: {health['valid_zip']}")
    if health["error"]:
        print(f"  ZIP error: {health['error']}")
    if health["members"]:
        print(f"  Members: {len(health['members'])}")
        for key in ("xl/workbook.xml", "xl/sharedStrings.xml"):
            print(f"    {key}: {'yes' if key in health['members'] else 'NO'}")

    methods = []

    # Method A: openpyxl
    for ro in (False, True):
        label = "openpyxl read_only" if ro else "openpyxl"
        try:
            wb = try_openpyxl(src, read_only=ro)
            copy_workbook(wb, out)
            print(f"\nOK: Recovered with {label} -> {out}")
            sys.exit(0)
        except Exception as e:
            methods.append(f"{label}: {e}")

    # Method B: XML parse from zip
    wb = recover_via_xml(src)
    if wb:
        wb.save(out)
        wb.close()
        print(f"\nOK: Recovered via XML parse -> {out}")
        print("  Review sheet names and row counts before replacing production file.")
        sys.exit(0)
    methods.append("xml parse: no readable worksheets")

    # Method C: repack zip then openpyxl
    tmp = out + ".repack.tmp.xlsx"
    if repack_zip(src, tmp):
        try:
            wb = load_workbook(tmp, data_only=True)
            copy_workbook(wb, out)
            os.remove(tmp)
            print(f"\nOK: Recovered via zip repack -> {out}")
            sys.exit(0)
        except Exception as e:
            methods.append(f"zip repack: {e}")
        finally:
            if os.path.isfile(tmp):
                os.remove(tmp)
    else:
        methods.append("zip repack: could not read zip members")

    print("\nFAILED: automatic recovery did not succeed.")
    print("Tried:")
    for m in methods:
        print(f"  - {m}")
    print("\nManual options:")
    print("  1. Excel: 文件 -> 打开 -> 选中文件旁下拉选「打开并修复」")
    print("  2. LibreOffice Calc: 打开损坏文件，另存为新 .xlsx")
    print("  3. 7-Zip: 右键 -> 解压；若部分 xml 能解压，可发给开发人员继续修复")
    print("  4. 检查服务器是否还有更早的 .corrupt.* 或管理后台下载过的备份")
    sys.exit(1)


if __name__ == "__main__":
    main()
