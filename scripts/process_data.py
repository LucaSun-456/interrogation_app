import json
import os
import sys
from openpyxl import load_workbook, Workbook

def process_data():
    input_file = os.path.join('data', 'experiment_data.xlsx')
    output_file = 'processed_data.xlsx'
    
    if not os.path.exists(input_file):
        print(f"找不到数据文件: {input_file}")
        sys.exit(1)
        
    print(f"正在读取数据文件: {input_file} ...")
    
    try:
        wb = load_workbook(input_file, data_only=True)
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        sys.exit(1)
        
    if 'participants' not in wb.sheetnames or 'interview_questionnaires' not in wb.sheetnames:
        print("数据文件中缺少必要的sheet (participants 或 interview_questionnaires)")
        sys.exit(1)
        
    # 读取 participants
    ws_p = wb['participants']
    p_headers = [c.value for c in ws_p[1]]
    participants = []
    for row in ws_p.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            participants.append(dict(zip(p_headers, row)))
            
    # 读取 interview_questionnaires
    ws_q = wb['interview_questionnaires']
    q_headers = [c.value for c in ws_q[1]]
    questionnaires = []
    for row in ws_q.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            questionnaires.append(dict(zip(q_headers, row)))
            
    # 按 group_name 分组
    groups = {}
    for p in participants:
        g = p.get('group_name')
        if g is not None:
            if g not in groups:
                groups[g] = []
            groups[g].append(p)
            
    results = []
    all_q_keys = set()
    
    for g, group_participants in groups.items():
        i_row = next((p for p in group_participants if p.get('role') == 'I'), None)
        s_row = next((p for p in group_participants if p.get('role') == 'S'), None)
        
        if not i_row or not s_row:
            continue
            
        i_phone = i_row.get('phone')
        s_phone = s_row.get('phone')
        
        s_guilt_raw = s_row.get('guilt')
        s_guilt = '有罪' if s_guilt_raw == 'Guilty' else ('无罪' if s_guilt_raw == 'Innocent' else s_guilt_raw)
        
        row_data = {
            '组别编号': g,
            '审讯者的组别': i_row.get('training_type'),
            '案件类型': s_row.get('case_type'),
            '嫌疑人是否有罪': s_guilt
        }
        
        # 查找问卷
        group_q = [q for q in questionnaires if q.get('phone') in (i_phone, s_phone)]
        
        i_post_guilt = None
        
        for q_row in group_q:
            role = q_row.get('role')
            phase = q_row.get('phase')
            
            role_cn = "审讯者" if role == "I" else "嫌疑人"
            phase_cn = "前测" if phase == "pre" else "后测"
            prefix = f"{role_cn}{phase_cn}_"
            
            ans_json = q_row.get('answers_json')
            if ans_json:
                try:
                    answers = json.loads(ans_json)
                    for k, v in answers.items():
                        col_name = f"{prefix}{k}"
                        row_data[col_name] = v
                        all_q_keys.add(col_name)
                        
                    if role == 'I' and phase == 'post':
                        i_post_guilt = answers.get('post_i_guilt')
                except:
                    pass
                    
        # 判断正确性
        if i_post_guilt and s_guilt_raw:
            is_correct = "未知"
            if i_post_guilt == "有罪" and s_guilt_raw == "Guilty":
                is_correct = "正确"
            elif i_post_guilt == "无罪" and s_guilt_raw == "Innocent":
                is_correct = "正确"
            elif i_post_guilt == "有罪" and s_guilt_raw == "Innocent":
                is_correct = "错误"
            elif i_post_guilt == "无罪" and s_guilt_raw == "Guilty":
                is_correct = "错误"
            
            row_data['审讯者的最终判断是否正确'] = is_correct
        else:
            row_data['审讯者的最终判断是否正确'] = ""
            
        results.append(row_data)
        
    if not results:
        print("没有找到任何完整的组别数据。")
        sys.exit(0)
        
    # 写入新Excel
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Processed Data"
    
    basic_cols = ['组别编号', '审讯者的组别', '案件类型', '嫌疑人是否有罪', '审讯者的最终判断是否正确']
    other_cols = sorted(list(all_q_keys))
    final_cols = basic_cols + other_cols
    
    # 写入表头
    out_ws.append(final_cols)
    
    # 写入数据
    for row_data in results:
        row = [row_data.get(c, "") for c in final_cols]
        out_ws.append(row)
        
    try:
        out_wb.save(output_file)
        print(f"数据处理成功！共处理了 {len(results)} 个组别。")
        print(f"结果已保存至: {os.path.abspath(output_file)}")
    except Exception as e:
        print(f"保存Excel文件失败: {e}")
        print("请确保没有在Excel中打开该文件。")

if __name__ == '__main__':
    process_data()
